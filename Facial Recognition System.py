import customtkinter as ctk
import face_recognition
import cv2
import numpy as np
from openpyxl import Workbook, load_workbook
from datetime import datetime
from tkinter import simpledialog, messagebox
import os
import time

# Initialize the known face data
known_face_encodings = []
known_face_names = []

# Load known faces
def load_known_faces():
    global known_face_encodings, known_face_names
    folder = "Students"
    for filename in os.listdir(folder):
        if filename.endswith((".jpeg", ".jpg", ".png")):
            image_path = os.path.join(folder, filename)
            image = face_recognition.load_image_file(image_path)
            encodings = face_recognition.face_encodings(image)
            if encodings:
                encoding = encodings[0]
                name = os.path.splitext(filename)[0]
                known_face_encodings.append(encoding)
                known_face_names.append(name)
            else:
                print(f"No faces found in {filename}. This image will be skipped.")

# Register a new student
def register_student():
    name = simpledialog.askstring("Register New Student", "Enter the student's name:")
    if name:
        video_capture = cv2.VideoCapture(0)
        if not video_capture.isOpened():
            messagebox.showerror("Error", "Unable to access the camera.")
            return

        messagebox.showinfo("Capture", "Press 'q' to capture the image of the student.")
        while True:
            ret, frame = video_capture.read()
            if not ret:
                messagebox.showerror("Error", "Failed to capture image.")
                video_capture.release()
                cv2.destroyAllWindows()
                return

            cv2.imshow("Capture Student Image", frame)
            if cv2.waitKey(1) & 0xFF == ord('q'):
                file_path = f"Students/{name}.jpeg"
                cv2.imwrite(file_path, frame)
                break

        video_capture.release()
        cv2.destroyAllWindows()
        
        image = face_recognition.load_image_file(file_path)
        encodings = face_recognition.face_encodings(image)
        
        if encodings:
            known_face_encodings.append(encodings[0])
            known_face_names.append(name)
            messagebox.showinfo("Success", f"{name} has been registered successfully!")
        else:
            os.remove(file_path)
            messagebox.showerror("Error", "No face detected in the image. Please try again.")
    else:
        messagebox.showwarning("Cancelled", "Registration cancelled.")

# Start attendance process
def start_attendance():
    print("Starting attendance process...")
    video_capture = cv2.VideoCapture(0)
    if not video_capture.isOpened():
        messagebox.showerror("Error", "Unable to access the camera.")
        return

    current_date = datetime.now().strftime("%d-%B")
    folder = "Attendance"
    if not os.path.exists(folder):
        os.makedirs(folder)
    
    file_name = os.path.join(folder, f"{current_date}.xlsx")
    
    try:
        workbook = load_workbook(file_name)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Name", "Time"])

    students = known_face_names.copy()
    tolerance = 0.5  # Adjusted for balanced accuracy

    while True:
        ret, frame = video_capture.read()
        if not ret:
            print("Failed to capture frame from camera.")
            continue

        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(rgb_frame)
        face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)

        for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
            face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
            best_match_index = np.argmin(face_distances)

            if face_distances[best_match_index] < tolerance:
                name = known_face_names[best_match_index]
            else:
                name = "Unknown"

            if name in students and name != "Unknown":
                students.remove(name)
                current_time = datetime.now().strftime("%H:%M:%S")
                sheet.append([name, current_time])
                print(f"Recorded {name} at {current_time}.")

            cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
            cv2.putText(frame, name, (left + 6, bottom - 6), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (255, 255, 255), 2)

        workbook.save(file_name)
        cv2.imshow("Camera", frame)

        if cv2.waitKey(1) & 0xFF == ord("q"):
                break

        
    if students:
        sheet.append([])
        sheet.append(["Absent Students", ""])
        for student in students:
            sheet.append([student, "Absent"])
        workbook.save(file_name)


    video_capture.release()
    cv2.destroyAllWindows()
    print("Attendance process ended.")

# View attendance
def view_attendance():
    folder = "Attendance"
    current_date = datetime.now().strftime("%d-%B")
    file_name = os.path.join(folder, f"{current_date}.xlsx")
    if os.path.exists(file_name):
        os.system(f"open {file_name}")
    else:
        messagebox.showerror("Error", f"No attendance file found for {current_date}.")

# Function to update title color based on theme
def update_title_color(label):
    current_mode = ctk.get_appearance_mode()
    if current_mode == "Dark":
        label.configure(text_color="white")
    else:
        label.configure(text_color="navy")

# Create the GUI with the restored color theme
def create_gui():
    ctk.set_appearance_mode("system")
    ctk.set_default_color_theme("blue")

    root = ctk.CTk()
    root.title("Ingenious Geniuses Attendance System")
    root.geometry("650x550")

    title_font = ("Trebuchet MS", 24, "bold")
    button_font = ("Verdana", 13)

    title_label = ctk.CTkLabel(root, text="Ingenious Geniuses Attendance System", font=title_font)
    title_label.pack(pady=(30, 10))

    update_title_color(title_label)

    button_style = {
        "width": 280,
        "height": 50,
        "corner_radius": 20,
        "font": button_font
    }

    ctk.CTkButton(root, text="Register New Student", command=register_student, fg_color="#3498db", **button_style).pack(pady=15)
    ctk.CTkButton(root, text="Start Attendance", command=start_attendance, fg_color="#2ecc71", **button_style).pack(pady=15)
    ctk.CTkButton(root, text="View Attendance", command=view_attendance, fg_color="#9b59b6", **button_style).pack(pady=15)
    ctk.CTkButton(root, text="Close Software", command=root.quit, fg_color="#e74c3c", **button_style).pack(pady=15)

    root.mainloop()

# Load known faces and run the GUI
load_known_faces()
create_gui()
